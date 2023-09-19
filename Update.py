from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

import re
import time
import asyncio
from datetime import datetime, timedelta
import requests
import openpyxl
from bs4 import BeautifulSoup

chrome_options = Options()
chrome_options.add_experimental_option("detach",True)
chrome_options.add_experimental_option("excludeSwitches",["enable-logging"])
user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
chrome_options.add_argument('user-agent=' + user_agent)
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)



# driver.implicitly_wait(5)


### 엑셀파일에 저장된 최근에 접속가능했던 주소를 가져와서 진행한다.
fpath = r'C:\Users\jjomi\PycharmProjects\NewtokkiCrawling\src\Newtoki.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['Site Information']
url = ws['A2'].value

# 프로그램 총 실행시간 체크
s_time = time.time()

# 만약 그 새 주소가 바뀌었다면, 숫자부분에 1씩 더하면서 바뀐 주소를 탐색해본다
while(1):
    try:
        temp = "/webtoon?toon=%EC%9D%BC%EB%B0%98%EC%9B%B9%ED%88%B0"
        url2 = url + temp
        driver.implicitly_wait(1)
        driver.get(url2)
        driver.find_element(By.XPATH,'//*[@id="content_wrapper"]/div[2]/div/section/div[1]/form/table/tbody/tr[2]/td/span[2]').click()
        driver.find_element(By.XPATH,'//*[@id="content_wrapper"]/div[2]/div/section/div[1]/form/table/tbody/tr[1]/td[2]/button').send_keys(Keys.ENTER)
        break
    except Exception as error:
        print(error)
        num = int(re.sub(r'[^0-9]', '', url)) + 1
        ### 여기 좀 수정해야할듯??
        url = "https://newtoki" + str(num) + ".com"


# 바뀐 주소 저장
k = driver.current_url.find('.com')
url3 = url2[:k+4]
# print(url3)
if url != url3:
    ws['A2'] = url3
    wb.save(fpath)

# 저장되 있는 마지막 업데이트 날짜를 기준으로 얼마나 업데이트할지 정하자.
last_update = ws['B2'].value
updating_date = datetime.now()
area = (updating_date-last_update).days

response = requests.get(driver.current_url,headers={'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'} )
html = response.text
soup = BeautifulSoup(html,'html.parser')
webtoon_list = soup.find_all('li',{'data-weekday':"월요일"})
for webtoon in webtoon_list:
    print(webtoon['data-initial'])
    print(webtoon['data-genre'])
    print(webtoon['date-title'])
    print("+++++++++++++++++++++++")

driver.find_element(By.XPATH, '//*[@id="fboardlist"]/div[4]/ul/li[4]/a').click()
response = requests.get(driver.current_url,headers={'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'} )
html = response.text
soup = BeautifulSoup(html,'html.parser')
webtoon_list = soup.find_all('li',{'data-weekday':"월요일"})
for webtoon in webtoon_list:
    print(webtoon['data-initial'])
    print(webtoon['data-genre'])
    print(webtoon['date-title'])
    print("+++++++++++++++++++++++")






###
### 로그인은 굳이 할 필요가 없을 것 같고, 캡챠입력은 수동으로 받도록 하자.
### https://www.youtube.com/watch?v=ZNOiwvrS_dc

### 걍 1화 링크 걸어주자

# 업데이트완료 날짜 저장
ws['B2'] = datetime.now()
wb.save(fpath)

driver.implicitly_wait(3)
driver.quit()
e_time = time.time()
print("총 실행시간 : " + str(e_time-s_time))