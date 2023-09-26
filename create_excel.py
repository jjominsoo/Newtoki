import openpyxl
import csv
import pandas
from datetime import datetime


wb = openpyxl.Workbook()
info = wb.create_sheet('Site Information')
info['A1'] = 'URL'
info['A2'] = "https://newtoki306.com"
info['B1'] = 'Last Update'
info['B2'] = datetime.now()
info['C1'] = 'Update List'
info['D1'] = 'Recent Webtoon Number'
info['D2'] = 2

week = wb.create_sheet('webtoon')
week['A1'] = '이름'
week['B1'] = '이미지'
week['C1'] = '추천수'
week['D1'] = '별점수'
week['E1'] = '총화수'
week['F1'] = '판타지'
week['G1'] = '액션'
week['H1'] = '개그'
week['I1'] = '미스터리'
week['J1'] = '로맨스'
week['K1'] = '드라마'
week['L1'] = '무협'
week['M1'] = '스포츠'
week['N1'] = '일상'
week['O1'] = '학원'
week['P1'] = '성인'
week['Q1'] = '한국'
week['R1'] = '중국'
week['S1'] = '요일'
week['T1'] = '1화링크'
week['U1'] = '상세페이지 URL'


wb.save(r'C:\Users\jjomi\PycharmProjects\NewtokkiCrawling\src\Newtoki6.xlsx')

# df = pandas.DataFrame(columns=['url'])
# df.to_csv("src/Newtoki.csv",index=False, encoding='UTF8')
# c = open('src/Newtoki.csv','a',newline='',encoding='UTF8')
# wc = csv.writer(c)
# wc.writerow(['https://newtoki307.com/webtoon/266771/%EB%AC%B5%ED%96%A5-%EB%8B%A4%ED%81%AC%EB%A0%88%EC%9D%B4%EB%94%94?sst=as_update&sod=desc&yoil=%EC%9B%94&toon=%EC%9D%BC%EB%B0%98%EC%9B%B9%ED%88%B0'])
# wc.writerow(['https://newtoki307.com/webtoon/11355191/%EB%8F%85%EB%B3%B4%EC%86%8C%EC%9A%94?sst=as_update&sod=desc&yoil=%EC%9B%94&toon=%EC%9D%BC%EB%B0%98%EC%9B%B9%ED%88%B0'])
# wc.writerow(['https://newtoki307.com/webtoon/26028719/%EC%95%84-%ED%98%95%EC%82%B0%ED%8C%8C?sst=as_update&sod=desc&yoil=%EC%9B%94&toon=%EC%9D%BC%EB%B0%98%EC%9B%B9%ED%88%B0'])