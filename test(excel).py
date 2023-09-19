# import time
# import requests
# from bs4 import BeautifulSoup
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
#
# from webdriver_manager.chrome import ChromeDriverManager
#
# chrome_options = Options()
# chrome_options.add_experimental_option("detach",True)
# chrome_options.add_experimental_option("excludeSwitches",["enable-logging"])
# chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36")
# service = Service(executable_path=ChromeDriverManager().install())
#
# # s_time = time.time()
# # driver = webdriver.Chrome(service=service, options=chrome_options)
# # url = "https://newtoki302.com/"
# # driver.get(url)
# # driver.find_element(By.XPATH, '//*[@id="sticky-wrapper"]/header/nav/a').send_keys(Keys.ENTER)
# # driver.find_element(By.XPATH, '//*[@id="sticky-wrapper"]/section/ul/li[2]/a').send_keys(Keys.ENTER)
# # driver.find_element(By.XPATH, '//*[@id="sticky-wrapper"]/section/ul/li[2]/ul/li[1]/a').send_keys(Keys.ENTER)
# # html = driver.page_source
# # soup = BeautifulSoup(html, 'html.parser')
# # print(html)
# # e_time = time.time()
# # print("Execution time = " + e_time-s_time)
# # driver.quit()
#
# # data = soup.select("div#root > div#wrap > div#container > div.content_wrap > div#content > div.component_wrap > ul")
# # data = soup.select('ul.ContentList__content_list--q5KXY')
# # print(data)
# # print(soup)
# # word = soup.select('.Poster__link--sopnC')
# # print(word)
# s_time2 = time.time()
# driver = webdriver.Chrome(service=service, options=chrome_options)
# url2 = "https://newtoki302.com/webtoon?toon=%EC%9D%BC%EB%B0%98%EC%9B%B9%ED%88%B0"
# driver.get(url2)
# driver.find_element(By.XPATH,'//*[@id="content_wrapper"]/div[2]/div/section/div[1]/form/table/tbody/tr[2]/td/span[2]').click()
# driver.find_element(By.XPATH,'//*[@id="content_wrapper"]/div[2]/div/section/div[1]/form/table/tbody/tr[1]/td[2]/button').send_keys(Keys.ENTER)
# webtoon_list = driver.find_elements(By.XPATH,'//*[@id="webtoon-list-all"]/li')
# for webtoon in webtoon_list:
#     print(webtoon.get_attribute('data-initial'))
#     print(webtoon.get_attribute('data-genre'))
#     print(webtoon.get_attribute('date-title'))
#     print("=====")
#
# driver.quit()
# e_time2 = time.time()
# print("Execution time = " + str(e_time2-s_time2))
import openpyxl
from datetime import datetime, timedelta

####
# fpath = r'C:\Users\jjomi\PycharmProjects\NewtokkiCrawling\src\temp.xlsx'
# wb = openpyxl.load_workbook(fpath)
#
# def work():
#     ws = wb['webtoon']
#     temp = ['로맨스','미스터리']
#
#     for l in temp:
#         # l = 만화의 장르
#         for i in ws['F1':'R1']:
#             # i = 저장할 장르 갯수 (13)
#             for j in i:
#                 # j = 장르
#                 print(j.column)
#                 if j.value == l:
#                     ws.cell(2, j.column, 1)
#                     print('Hello\t' + j.value)
#                 else:
#                     ws.cell(2,j.column, 0)
#                     print("Bye\t" + j.value)
#
# def format():
#     wb.remove(wb['webtoon'])
#     week = wb.create_sheet('webtoon')
#     week['A1'] = '이름'
#     week['B1'] = '이미지'
#     week['C1'] = '추천수'
#     week['D1'] = '별점수'
#     week['E1'] = '총화수'
#     week['F1'] = '판타지'
#     week['G1'] = '액션'
#     week['H1'] = '개그'
#     week['I1'] = '미스터리'
#     week['J1'] = '로맨스'
#     week['K1'] = '드라마'
#     week['L1'] = '무협'
#     week['M1'] = '스포츠'
#     week['N1'] = '일상'
#     week['O1'] = '학원'
#     week['P1'] = '성인'
#     week['Q1'] = '한국'
#     week['R1'] = '중국'
#     week['S1'] = '요일'
#     week['T1'] = '1화링크'
#     week['U1'] = '최신화링크'
# format()
# work()
####
import pandas
import csv
import os
from openpyxl.drawing.image import Image
from openpyxl import Workbook
fpath = r'C:\Users\jjomi\PycharmProjects\NewtokkiCrawling\src\Newtoki3.xlsx'
wb = openpyxl.load_workbook(fpath)
wb_siteInfo = wb['Site Information']
wb_webtoon = wb['webtoon']
df = pandas.read_csv('src/Newtoki.csv')

# img_dir = 'src/img/'
# no_img_dir = 'src/files/'
# img_file_list = sorted(os.listdir(img_dir),key=lambda x:int(x.split('.')[0]))
# for i, img_file in enumerate(img_file_list):
#     try:
#         image_path = os.path.join(img_dir, img_file)
#         image = Image(image_path)
#         wb_webtoon.add_image(image, anchor='B' + str(i + 2))
#     except Exception as e:
#         print(e)
#         image_path = os.path.join(no_img_dir, 'no_image.png')
#         image = Image(image_path)
#         wb_webtoon.add_image(image, anchor='B' + str(i + 2))
#
#     if i == 0:
#         wb_webtoon.column_dimensions['B'].width = image.width*63.2/504.19
#     wb_webtoon.row_dimensions[i+2].height = image.height*225.35/298.96
print(df['제목'])


wb.save(fpath)

# (updating_date-last_update).days